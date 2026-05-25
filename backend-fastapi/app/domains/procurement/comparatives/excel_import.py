from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any, Optional

import openpyxl
from openpyxl.utils.datetime import from_excel

from app.ai.ai_client import normalize_comparative_json


_STOP_WORDS = (
    "total ofertado",
    "total ofertas",
    "oferta homogenea",
    "forma de pago",
    "observaciones",
    "garant",
    "retencion",
    "plazo",
    "hito",
    "especificaciones",
)


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _normalize(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(char for char in normalized if not unicodedata.combining(char)).lower()


def _as_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        # Limpia monedas/simbolos comunes manteniendo solo digitos y separadores.
        text = (
            text.replace("\xa0", "")
            .replace("€", "")
            .replace("EUR", "")
            .replace("eur", "")
            .replace("%", "")
            .replace(" ", "")
        )
        text = re.sub(r"[^0-9,.\-]", "", text)
        if not text:
            return None
        # Detecta separador decimal por ultima ocurrencia cuando hay ambos.
        if "," in text and "." in text:
            last_comma = text.rfind(",")
            last_dot = text.rfind(".")
            if last_comma > last_dot:
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _looks_like_label(text: str) -> bool:
    lower = _normalize(text)
    exact_labels = {
        "precio",
        "importe",
        "precios",
        "datos",
        "datos oferta",
        "datos de oferta",
        "porcentajes",
        "planificacion",
    }
    if lower in exact_labels:
        return True
    return any(
        token in lower
        for token in (
            "precios",
            "datos",
            "planificacion",
            "porcentajes",
            "tlf",
        )
    )


def _cell_text_with_merged(sheet, row: int, col: int) -> str:
    text = _as_text(sheet.cell(row=row, column=col).value)
    if text:
        return text
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return _as_text(sheet.cell(row=merged.min_row, column=merged.min_col).value)
    return ""


def _clean_provider_name(value: str) -> Optional[str]:
    if not value:
        return None
    text = re.sub(r"\s+", " ", value).strip(" .:-")
    # Limpia prefijos de cabecera comunes: "OFERTA 1 - PROVEEDOR X"
    text = re.sub(r"^\s*OFERTA\s*\d+\s*[-:./]?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^\s*(DATOS\s+DE\s+)?OFERTA\s*[-:./]?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^\s*(PRECIO|IMPORTE)\s*[-:./]?\s*", "", text, flags=re.IGNORECASE)
    text = text.strip(" .:-")
    if not text:
        return None
    if _looks_like_label(text):
        return None
    if len(text) < 3:
        return None
    return text


def _extract_jefe_obra_hint(sheet) -> Optional[str]:
    for row in range(1, min(sheet.max_row, 20) + 1):
        for col in range(1, min(sheet.max_column, 60) + 1):
            text = _as_text(sheet.cell(row=row, column=col).value)
            if not text:
                continue
            lower = _normalize(text)
            if "jefe" in lower and "obra" in lower:
                value = _find_first_value_right(sheet, row, col)
                if value is None:
                    continue
                jefe = _clean_provider_name(_as_text(value))
                if jefe:
                    return jefe
    return None


def _find_first_value_right(sheet, row: int, col: int, max_steps: int = 6) -> Optional[Any]:
    for offset in range(1, max_steps + 1):
        cell = sheet.cell(row=row, column=col + offset)
        if cell.value is not None and str(cell.value).strip() != "":
            return cell.value
    return None


def _find_first_cell_right(sheet, row: int, col: int, max_steps: int = 6):
    for offset in range(1, max_steps + 1):
        cell = sheet.cell(row=row, column=col + offset)
        if cell.value is not None and str(cell.value).strip() != "":
            return cell
    return None


def _find_header_row(sheet) -> int:
    max_row = min(sheet.max_row, 120)
    max_col = min(sheet.max_column, 80)
    for row in range(1, max_row + 1):
        texts = [
            _normalize(_as_text(sheet.cell(row=row, column=col).value))
            for col in range(1, max_col + 1)
        ]
        has_desc = any(
            ("descrip" in text) or ("concept" in text) or ("partida" in text)
            for text in texts
        )
        has_code = any(
            text.startswith("cod") or ("codigo" in text) or ("referen" in text) or (text == "ref")
            for text in texts
        )
        has_measure = any(
            ("med" in text) or ("medicion" in text) or (text == "ud") or ("unidad" in text)
            for text in texts
        )
        if has_desc and (has_code or has_measure):
            return row
    raise ValueError("No se encontro la fila de cabecera (Cod./Descripcion).")


def _find_price_row(sheet, start_row: int) -> int:
    for row in range(start_row, min(sheet.max_row, start_row + 12) + 1):
        texts = [_normalize(_as_text(sheet.cell(row=row, column=col).value)) for col in range(1, 60)]
        has_price = any("precio" in text for text in texts)
        has_importe = any("import" in text for text in texts)
        if has_price and has_importe:
            return row
    raise ValueError("No se encontro la fila de precios (Precio/Importe).")


def _find_column_by_label(sheet, row: int, label: str) -> Optional[int]:
    label_lower = _normalize(label)
    for col in range(1, 30):
        text = _normalize(_as_text(sheet.cell(row=row, column=col).value))
        if label_lower in text:
            return col
    return None


def _detect_provider(sheet, row_from: int, col: int, min_row: int = 1) -> Optional[str]:
    for row in range(row_from, max(min_row, row_from - 8), -1):
        text = _cell_text_with_merged(sheet, row=row, col=col)
        if not text:
            continue
        cleaned = _clean_provider_name(text)
        if not cleaned:
            continue
        return cleaned
    return None


def _extract_date(sheet, row: int, col: int) -> Optional[str]:
    cell = sheet.cell(row=row, column=col)
    value = cell.value
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.date().isoformat()
    if isinstance(value, (int, float)) and cell.is_date:
        return from_excel(value).date().isoformat()
    return None


def parse_comparative_excel(content: bytes, filename: str | None = None) -> dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = wb.active
    header_row = _find_header_row(sheet)
    price_row = _find_price_row(sheet, header_row + 1)

    cod_col = (
        _find_column_by_label(sheet, header_row, "cod")
        or _find_column_by_label(sheet, header_row, "referen")
        or _find_column_by_label(sheet, header_row, "ref")
        or 3
    )
    med_col = _find_column_by_label(sheet, header_row, "med") or 4
    ud_col = _find_column_by_label(sheet, header_row, "ud") or 5
    desc_col = (
        _find_column_by_label(sheet, header_row, "descrip")
        or _find_column_by_label(sheet, header_row, "concept")
        or _find_column_by_label(sheet, header_row, "partida")
        or 6
    )
    jefe_obra_hint = _extract_jefe_obra_hint(sheet)
    jefe_obra_key = _normalize(jefe_obra_hint) if jefe_obra_hint else None

    provider_columns: list[tuple[str, int, int]] = []
    seen_providers: set[str] = set()
    for col in range(1, sheet.max_column + 1):
        text = _normalize(_as_text(sheet.cell(row=price_row, column=col).value))
        if "precio" not in text:
            continue
        provider = _clean_provider_name(_cell_text_with_merged(sheet, header_row, col))
        if not provider:
            provider = _clean_provider_name(_cell_text_with_merged(sheet, header_row, max(1, col - 1)))
        if not provider:
            provider = _detect_provider(sheet, header_row, col, min_row=max(1, header_row - 8))
        if not provider:
            provider = _detect_provider(sheet, header_row, max(1, col - 1), min_row=max(1, header_row - 8))
        if not provider:
            continue
        provider_key = provider.strip().lower()
        if jefe_obra_key and _normalize(provider) == jefe_obra_key:
            continue
        if provider_key in seen_providers:
            continue
        importe_col = col + 1
        importe_label = _normalize(_as_text(sheet.cell(row=price_row, column=importe_col).value))
        if "import" not in importe_label:
            continue
        seen_providers.add(provider_key)
        provider_columns.append((provider, col, importe_col))

    # Fallback: algunos templates no etiquetan "precio" y solo ponen "importe".
    if not provider_columns:
        for col in range(2, sheet.max_column + 1):
            text = _normalize(_as_text(sheet.cell(row=price_row, column=col).value))
            if "import" not in text:
                continue
            price_col = col - 1
            provider = _clean_provider_name(_cell_text_with_merged(sheet, header_row, price_col))
            if not provider:
                provider = _clean_provider_name(_cell_text_with_merged(sheet, header_row, col))
            if not provider:
                provider = _detect_provider(sheet, header_row, price_col, min_row=max(1, header_row - 8))
            if not provider:
                provider = _detect_provider(sheet, header_row, col, min_row=max(1, header_row - 8))
            if not provider:
                continue
            provider_key = provider.strip().lower()
            if jefe_obra_key and _normalize(provider) == jefe_obra_key:
                continue
            if provider_key in seen_providers:
                continue
            seen_providers.add(provider_key)
            provider_columns.append((provider, price_col, col))

    if not provider_columns:
        raise ValueError("No se encontraron proveedores en el comparativo.")

    provider_totals: dict[str, float] = {}
    lines: list[dict[str, Any]] = []
    empty_rows = 0
    last_data_row = price_row
    for row in range(price_row + 1, sheet.max_row + 1):
        cod = _as_text(sheet.cell(row=row, column=cod_col).value)
        descripcion = _as_text(sheet.cell(row=row, column=desc_col).value)
        if descripcion and any(token in descripcion.lower() for token in _STOP_WORDS):
            last_data_row = row - 1
            break
        if not cod and not descripcion:
            empty_rows += 1
            if empty_rows >= 5:
                break
            continue
        empty_rows = 0

        cantidad = _as_number(sheet.cell(row=row, column=med_col).value)
        unidad = _as_text(sheet.cell(row=row, column=ud_col).value)

        prices = []
        for provider, price_col, importe_col in provider_columns:
            precio = _as_number(sheet.cell(row=row, column=price_col).value)
            importe = _as_number(sheet.cell(row=row, column=importe_col).value)
            if precio is None and importe is None:
                continue
            prices.append(
                {
                    "proveedor": provider,
                    "precio": precio,
                    "importe": importe,
                }
            )

        if not descripcion and not prices:
            continue

        lines.append(
            {
                "cod_capitulo": cod or None,
                "cantidad": cantidad,
                "unidad": unidad or None,
                "descripcion": descripcion or None,
                "prices": prices,
            }
        )

    totals_search_start = last_data_row + 1 if last_data_row else price_row + 1
    totals_search_end = min(sheet.max_row, totals_search_start + 30)
    max_col_for_totals = min(sheet.max_column, 80)
    for row in range(totals_search_start, totals_search_end + 1):
        row_texts = [
            _normalize(_as_text(sheet.cell(row=row, column=col).value))
            for col in range(1, max_col_for_totals + 1)
        ]
        joined = " ".join(row_texts)
        if "total ofertado" not in joined and "total ofertas" not in joined:
            continue
        for provider, _price_col, importe_col in provider_columns:
            amount = _as_number(sheet.cell(row=row, column=importe_col).value)
            if amount is None:
                amount = _as_number(sheet.cell(row=row, column=importe_col - 1).value)
            if amount is None or amount <= 0:
                continue
            key = provider.strip().lower()
            if key not in provider_totals or amount > provider_totals[key]:
                provider_totals[key] = amount

    header: dict[str, Any] = {}
    obra_num = None
    obra_nombre = None
    for row in range(1, min(sheet.max_row, 20) + 1):
        for col in range(1, min(sheet.max_column, 60) + 1):
            text = _as_text(sheet.cell(row=row, column=col).value)
            if not text:
                continue
            lower = _normalize(text)
            if lower.replace(":", "").strip() == "obra":
                value = _find_first_value_right(sheet, row, col)
                if value is not None:
                    obra_text = _as_text(value)
                    match = re.match(r"^(\d+)[-_\s]*(.*)$", obra_text)
                    if match:
                        obra_num = match.group(1)
                        obra_nombre = match.group(2) or obra_text
                    else:
                        obra_nombre = obra_text
            if "descripcion" in lower and ":" in text:
                value = _find_first_value_right(sheet, row, col)
                if value is not None:
                    header["descripcion"] = _as_text(value)
            if "jefe" in lower and "obra" in lower:
                value = _find_first_value_right(sheet, row, col)
                if value is not None:
                    header["jefe_obra"] = _as_text(value)
            if "planific" in lower:
                value = _find_first_value_right(sheet, row, col)
                if value is not None:
                    header["planificacion_num"] = _as_text(value)
            if "fecha solicitud" in lower:
                cell = _find_first_cell_right(sheet, row, col)
                if not cell:
                    continue
                date = _extract_date(sheet, cell.row, cell.column)
                if date:
                    header["fecha_solicitud"] = date
            if "fecha aprobacion" in lower:
                cell = _find_first_cell_right(sheet, row, col)
                if not cell:
                    continue
                date = _extract_date(sheet, cell.row, cell.column)
                if date:
                    header["fecha_aprobacion"] = date

    if obra_num:
        header["obra_num"] = obra_num
    if obra_nombre:
        header["obra_nombre"] = obra_nombre

    providers = [{"name": name, "offer_num": None} for name, _, _ in provider_columns]
    offers = [
        {
            "offer_name": name,
            "supplier_name": name,
            "total_amount": provider_totals.get(name.strip().lower()),
        }
        for name, _, _ in provider_columns
    ]

    raw = {
        "header": header,
        "providers": providers,
        "lines": lines,
    }
    normalized = normalize_comparative_json(raw) or {}
    if not isinstance(normalized, dict):
        raise ValueError("No se pudo normalizar el comparativo Excel.")
    normalized_header = normalized.get("header")
    if not isinstance(normalized_header, dict):
        normalized_header = {}
        normalized["header"] = normalized_header

    obra_num_norm = normalized_header.get("obra_num")
    obra_nombre_norm = normalized_header.get("obra_nombre")
    if not obra_num_norm and isinstance(obra_nombre_norm, str):
        match = re.match(r"^(\d+)[-_\s]*(.*)$", obra_nombre_norm.strip())
        if match:
            normalized_header["obra_num"] = match.group(1)
            normalized_header["obra_nombre"] = match.group(2) or obra_nombre_norm
    normalized["offers"] = offers
    normalized["source"] = "excel"
    if filename:
        normalized["source_filename"] = filename
    return normalized

